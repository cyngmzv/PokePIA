#En este script se encuentran todas las funciones que se utilizan para guardar los datos en excel y archivo de texto.
from openpyxl import Workbook, load_workbook
import requests
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
import os
import json
# Función para actualizar datos de un pokemon en un archivo Excel
def actualizar_datos_en_excel(datos, nombre_archivo):
    if not os.path.exists(nombre_archivo):
        # Si el archivo no existe, crea uno nuevo
        libro = Workbook()
        hoja = libro.active
        hoja.title = 'Datos'
        hoja.append(['Nombre', 'Tipo', 'HP', 'Ataque', 'Defensa', 'Peso'])
        libro.save(nombre_archivo)
    else:
        # Si el archivo ya existe, carga el libro existente
        libro = load_workbook(nombre_archivo)
        hoja = libro.active if 'Datos' in libro.sheetnames else libro.create_sheet(title='Datos')

    for pokemon in datos:
        hoja.append([
            pokemon['name'],
            ', '.join(pokemon.get('type', ['Desconocido'])),  # Convertir lista a cadena
            pokemon.get('hp', 0),
            pokemon.get('attack', 0),
            pokemon.get('defense', 0),
            pokemon.get('weight', 0)
        ])

    libro.save(nombre_archivo)


# Función para actualizar datos de un ítem en un archivo Excel
def actualizar_datos_item_en_excel(datos, nombre_archivo):
    if not os.path.exists(nombre_archivo):
        # Si el archivo no existe, crea uno nuevo
        libro = Workbook()
        hoja = libro.active
        hoja.title = 'Datos'
        hoja.append(['Nombre', 'ID', 'Costo', 'Fling Power', 'Efecto'])
        libro.save(nombre_archivo)
    else:
        # Si el archivo ya existe, carga el libro existente
        libro = load_workbook(nombre_archivo)
        hoja = libro.active if 'Datos' in libro.sheetnames else libro.create_sheet(title='Datos')

    for item in datos:
        hoja.append([item['name'], item['id'], item['cost'], item['fling_power'], item['effect']])

    libro.save(nombre_archivo)

# Función para actualizar datos de evoluciones en un archivo Excel
def actualizar_datos_evoluciones_en_excel(datos, nombre_archivo):
    if not os.path.exists(nombre_archivo):
        # Si el archivo no existe, crea uno nuevo
        libro = Workbook()
        hoja = libro.active
        hoja.title = 'Evoluciones'
        hoja.append(['Nombre', 'Evoluciona a'])
        libro.save(nombre_archivo)
    else:
        # Si el archivo ya existe, carga el libro existente
        libro = load_workbook(nombre_archivo)
        hoja = libro.active if 'Evoluciones' in libro.sheetnames else libro.create_sheet(title='Evoluciones')

    for evolucion in datos:
        hoja.append([evolucion['name'], evolucion['evoluciona a']])

    libro.save(nombre_archivo)

# Función para guardar datos en un archivo TXT
def guardar_datos_en_txt(datos, nombre_archivo):
    with open(nombre_archivo, 'a') as archivo:
        for item in datos:
            archivo.write(f"{json.dumps(item, indent=2)}\n")

def guardar_promedio_stats_en_excel(promedio_stats, nombre_archivo):
    if not os.path.exists(nombre_archivo):
        # Si el archivo no existe, crea uno nuevo
        libro = Workbook()
        hoja = libro.active
        hoja.title = 'Promedio Stats'
        hoja.append(['Promedio HP', 'Promedio Ataque', 'Promedio Defensa'])
        libro.save(nombre_archivo)
    else:
        # Si el archivo ya existe, carga el libro existente
        libro = load_workbook(nombre_archivo)
        hoja = libro.active if 'Promedio Stats' in libro.sheetnames else libro.create_sheet(title='Promedio Stats')

    hoja.append([
        promedio_stats.get('Promedio HP', 0),
        promedio_stats.get('Promedio Ataque', 0),
        promedio_stats.get('Promedio Defensa', 0),
    ])

    libro.save(nombre_archivo)

def guardar_pokemon_mas_fuerte_en_excel(pokemon_fuerte, nombre_archivo):
    if not os.path.exists(nombre_archivo):
        # Si el archivo no existe, crea uno nuevo
        libro = Workbook()
        hoja = libro.active
        hoja.title = 'Pokemon Más Fuerte'
        hoja.append(['Pokemon más fuerte', 'Promedio de stats'])
        libro.save(nombre_archivo)
    else:
        # Si el archivo ya existe, carga el libro existente
        libro = load_workbook(nombre_archivo)
        hoja = libro.active if 'Pokemon Más Fuerte' in libro.sheetnames else libro.create_sheet(title='Pokemon Más Fuerte')

    hoja.append([
        pokemon_fuerte.get('Pokemon más fuerte', 'Desconocido'),
        pokemon_fuerte.get('Promedio de stats', 0),
    ])

    libro.save(nombre_archivo)

def guardar_pokemon_mas_debil_en_excel(pokemon_debil, nombre_archivo):
    if not os.path.exists(nombre_archivo):
        # Si el archivo no existe, crea uno nuevo
        libro = Workbook()
        hoja = libro.active
        hoja.title = 'Pokemon Más Débil'
        hoja.append(['Pokemon más débil', 'Promedio de stats'])
        libro.save(nombre_archivo)
    else:
        # Si el archivo ya existe, carga el libro existente
        libro = load_workbook(nombre_archivo)
        hoja = libro.active if 'Pokemon Más Débil' in libro.sheetnames else libro.create_sheet(title='Pokemon Más Débil')

    hoja.append([
        pokemon_debil.get('Pokemon más débil', 'Desconocido'),
        pokemon_debil.get('Promedio de stats', 0),
    ])

    libro.save(nombre_archivo)

# Función para guardar el ticket de compras en un archivo Excel
def guardar_ticket_en_excel(ticket, nombre_archivo):
    if not os.path.exists(nombre_archivo):
        # Si el archivo no existe, crea uno nuevo
        libro = Workbook()
        hoja = libro.active
        hoja.title = 'Ticket de Compras'
        hoja.append(['Ítem', 'Precio', 'Precio con IVA (16%)'])
        libro.save(nombre_archivo)
    else:
        # Si el archivo ya existe, carga el libro existente
        libro = load_workbook(nombre_archivo)
        hoja = libro.active if 'Ticket de Compras' in libro.sheetnames else libro.create_sheet(title='Ticket de Compras')

    # Añadir encabezado solo si la hoja está vacía
    if hoja.max_row == 1:
        hoja.append(['Ítem', 'Precio', 'Precio con IVA (16%)', 'Subtotal', 'Total'])

    for item in ticket:
        hoja.append([item['name'], item['cost'], item['cost_with_iva'], '', ''])

    # Calcular y agregar Subtotal y Total al final del ticket
    subtotal = sum(item['cost'] for item in ticket)
    total_con_iva = sum(item['cost_with_iva'] for item in ticket)
    hoja.append(['', '', '', subtotal, total_con_iva])

    libro.save(nombre_archivo)

# Función para guardar el ticket de compras en un archivo de texto
def guardar_ticket_en_txt(ticket, nombre_archivo):
    with open(nombre_archivo, 'w') as archivo:
        archivo.write("Ticket de Compras:\n")
        for item in ticket:
            archivo.write(f"{item['name']}: ${item['cost']} (IVA del 16%: ${item['cost_with_iva']})\n")

        subtotal = sum(item['cost'] for item in ticket)
        total_con_iva = sum(item['cost_with_iva'] for item in ticket)
        archivo.write(f"\nSubtotal: ${subtotal}\n")
        archivo.write(f"Total (con IVA del 16%): ${total_con_iva:.2f}\n")
