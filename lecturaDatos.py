from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

#función para leer archivs txt
def leertexto():
    nombre_archivo = input("Ingrese el nombre del archivo de texto (con extensión .txt): ")
    try:
        with open(nombre_archivo, 'r') as archivo:
            contenido = archivo.read()
            print(contenido)
    except Exception as e:
        print(f"Error al leer el archivo de texto: {e}")

#función para leer archivos xlsx
def leerexcel():
    while True:
        print("Ingrese el nombre de su archivo SIN la terminación '.xlsx'")
        print("Verificar que esté correctamente escrito, que el archivo se encuentre en esta carpeta y que los datos de su archivo estén en el rango de celdas A1 a D40")
        archivo_excel = str(input())
        try:
            libro = load_workbook(archivo_excel + '.xlsx')
        except FileNotFoundError:
            print("El archivo no existe")
            return
        except InvalidFileException:
            print("El archivo no es un archivo de Excel válido.")
            return

        print("Cantidad de hojas:", len(libro.sheetnames))
        numHoja = 0
        for nombreHoja in libro.sheetnames:
            print("Hoja", numHoja + 1, nombreHoja)
            if nombreHoja == "Alumnos":
                libro.active = numHoja
            numHoja += 1
        hoja = libro.active
        print("La página activa es:", hoja.title)
        
        # Una forma de leer
        rango = hoja["A1":"D40"]  # Adaptar rango a la totalidad de celdas con info
        for celda in rango:
            for objeto in celda:
                print(objeto, "contiene", objeto.value)
                
        # Leer por filas
        print("Impresión de todas las celdas")
        for fila in hoja.values:
            for valor in fila:
                print(valor, end="\t")
            print("\n")

        opcion = input("Presiona 'M' para regresar al menú principal o cualquier otra tecla para salir: ")
        if opcion.upper() == 'M':
            break

